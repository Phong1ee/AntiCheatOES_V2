"""Database-backed report exports delivered by the shared report worker."""

import io
import os
import tempfile
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from src.a_db_config import BackgroundJob, BackgroundJobStatus, BackgroundJobType, Exam
from src.service.background_job_service import complete_job, get_or_create_job, mark_job_running
from src.service.outbox_publisher import enqueue_outbox_event
from src.service.object_storage import ObjectNotFoundError, storage_for


REPORT_TYPE_EXAM_RESULTS = "exam_results"


def request_exam_results_report(
    db: Session,
    *,
    exam_id: int,
    requested_by: str,
    request_id: str,
) -> tuple[BackgroundJob, bool]:
    """Add one report job and its outbox event to the caller transaction."""
    job, duplicate = get_or_create_job(
        db,
        job_type=BackgroundJobType.report_export,
        requested_by=requested_by,
        business_key=f"report:{REPORT_TYPE_EXAM_RESULTS}:exam:{exam_id}:request:{request_id}",
        result_metadata={"report_type": REPORT_TYPE_EXAM_RESULTS, "exam_id": exam_id, "format": "xlsx"},
    )
    if not duplicate:
        enqueue_outbox_event(
            db,
            event_type="report.requested",
            aggregate_type="background_job",
            aggregate_id=job.job_id,
            metadata={"job_id": job.job_id, "report_type": REPORT_TYPE_EXAM_RESULTS, "exam_id": exam_id},
        )
    return job, duplicate


def report_job_summary(job: BackgroundJob) -> dict:
    metadata = job.result_metadata if isinstance(job.result_metadata, dict) else {}
    status = job.status.value if hasattr(job.status, "value") else str(job.status)
    return {
        "jobId": job.job_id,
        "status": status,
        "reportType": metadata.get("report_type"),
        "examId": metadata.get("exam_id"),
        "format": metadata.get("format"),
        "createdAt": job.created_at.isoformat() if job.created_at else None,
        "completedAt": job.completed_at.isoformat() if job.completed_at else None,
        "error": job.last_error if status == BackgroundJobStatus.failed.value else None,
    }


def _report_directory() -> Path:
    configured = os.getenv("REPORT_EXPORT_DIR", "generated_reports").strip()
    directory = Path(configured)
    if not directory.is_absolute():
        directory = Path(__file__).resolve().parents[2] / directory
    directory.mkdir(parents=True, exist_ok=True)
    return directory


_STRATEGY_LABELS = {
    "highest": "Highest score across attempts",
    "average": "Average across attempts",
    "last_attempt": "Latest submitted attempt",
}


def _exam_results_workbook(db: Session, exam: Exam) -> bytes:
    """Build the exam summary plus one row per student for their final-result attempt."""
    # Local import avoids a module cycle: resultsRoute imports this module at load time.
    from src.route.teacherRoute.resultsRoute import _build_student_rows, _exam_stats

    stats = _exam_stats(db, exam)
    students = _build_student_rows(db, exam)
    scale = stats["gradingScale"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Exam Results"

    bold = Font(bold=True)
    summary_rows = [
        ("Exam", exam.title),
        ("Grading Method", _STRATEGY_LABELS.get(stats["resultStrategy"], stats["resultStrategy"])),
        ("Average Score", f"{stats['avgScore']:.2f} / {scale:.0f}"),
        ("Highest Score", f"{stats['highestScore']:.2f} / {scale:.0f}"),
        ("Lowest Score", f"{stats['lowestScore']:.2f} / {scale:.0f}"),
        ("Students Submitted", f"{stats['submittedCount']} / {stats['totalStudents']}"),
    ]
    for label, value in summary_rows:
        sheet.append([label, value])
        sheet.cell(row=sheet.max_row, column=1).font = bold
    sheet.append([])

    header = ["Student ID", "Name", "Final Score (/100)", "Correct Answers", "Total Questions", "Time Spent", "Status", "Submitted At"]
    sheet.append(header)
    header_row = sheet.max_row
    for column in range(1, len(header) + 1):
        sheet.cell(row=header_row, column=column).font = bold

    for row in students:
        sheet.append([
            row["studentId"],
            row["name"],
            row["score"],
            row["correctAnswers"],
            row["totalQuestions"],
            row["timeSpent"],
            row["status"],
            row["submittedAt"] or "",
        ])

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
    for column_cells in sheet.columns:
        values = [str(cell.value) for cell in column_cells if cell.value is not None]
        width = max((len(value) for value in values), default=10)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(width + 2, 40)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _write_artifact(job_id: int, data: bytes) -> str:
    filename = f"report_job_{job_id}.xlsx"
    storage_for("reports", _report_directory()).put(filename, data)
    return filename


def report_artifact_path(job_id: int) -> Path:
    """Resolve a worker-created artifact without trusting a database path value."""
    return _report_directory() / f"report_job_{job_id}.xlsx"


def report_artifact_bytes(job: BackgroundJob) -> bytes:
    """Read a private report object using trusted job metadata only."""
    metadata = job.result_metadata if isinstance(job.result_metadata, dict) else {}
    key = metadata.get("artifact_key") or f"report_job_{job.job_id}.xlsx"
    if not isinstance(key, str):
        raise ObjectNotFoundError("invalid report artifact key")
    return storage_for("reports", _report_directory()).get(key)


def handle_report_requested(envelope: dict, db: Session) -> None:
    """Generate an export once; the worker's processed-event marker handles redelivery."""
    if envelope.get("event_type") != "report.requested":
        raise ValueError("report.queue accepts report.requested events only")
    try:
        job_id = int(envelope["aggregate_id"])
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError("report.requested requires a numeric background job aggregate_id") from exc

    job = db.get(BackgroundJob, job_id, with_for_update=True)
    if not job:
        raise ValueError("Report background job not found")
    job_type = job.job_type.value if hasattr(job.job_type, "value") else str(job.job_type)
    status = job.status.value if hasattr(job.status, "value") else str(job.status)
    if job_type != BackgroundJobType.report_export.value:
        raise ValueError("Background job is not a report export")
    if status == BackgroundJobStatus.completed.value:
        return
    if status == BackgroundJobStatus.failed.value:
        return

    metadata = job.result_metadata if isinstance(job.result_metadata, dict) else {}
    if metadata.get("report_type") != REPORT_TYPE_EXAM_RESULTS:
        raise ValueError("Unsupported report type")
    try:
        exam_id = int(metadata["exam_id"])
    except (KeyError, TypeError, ValueError) as exc:
        raise ValueError("Report job is missing its exam identifier") from exc
    exam = db.get(Exam, exam_id)
    if not exam:
        raise ValueError("Report exam no longer exists")

    mark_job_running(db, job_id)
    artifact_key = _write_artifact(job_id, _exam_results_workbook(db, exam))
    complete_job(
        db,
        job_id,
        processed_rows=1,
        success_rows=1,
        failed_rows=0,
        result_metadata={
            "report_type": REPORT_TYPE_EXAM_RESULTS,
            "exam_id": exam_id,
            "format": "xlsx",
            "artifact_key": artifact_key,
            "generated_at": datetime.now(timezone.utc).isoformat(),
        },
    )
