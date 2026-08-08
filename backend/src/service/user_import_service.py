"""XLSX-only parsing utilities for the non-persisting Admin user import preview."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook


REQUIRED_HEADERS = ("school_id", "full_name", "email", "role", "phone", "date_of_birth", "initial_password")
BLOCKED_HEADERS = {"id", "password_hash", "is_locked", "locked_at", "locked_by", "deleted_at", "deleted_by", "created_at", "updated_at"}
MAX_USER_ROWS = 1000


class UserImportParseError(ValueError):
    pass


@dataclass
class ParsedUserImportRow:
    row_number: int
    values: dict[str, Any]
    errors: list[str] = field(default_factory=list)


def _header(value: Any) -> str:
    return str(value or "").strip().casefold()


def _blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _date(value: Any) -> date | None:
    if _blank(value):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip())
        except ValueError as exc:
            raise ValueError("Date of birth must be an Excel date or YYYY-MM-DD") from exc
    raise ValueError("Date of birth must be an Excel date or YYYY-MM-DD")


def parse_user_import_xlsx(content: bytes) -> list[ParsedUserImportRow]:
    """Read the constrained workbook format; this function never touches the DB."""
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise UserImportParseError("The uploaded file is not a valid XLSX workbook") from exc
    if "Users" not in workbook.sheetnames:
        raise UserImportParseError("The workbook must contain a Users sheet")
    sheet = workbook["Users"]
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = next(rows)
    except StopIteration as exc:
        raise UserImportParseError("The Users sheet is missing its header row") from exc
    normalized_headers = [_header(value) for value in headers]
    if any(header in BLOCKED_HEADERS for header in normalized_headers):
        raise UserImportParseError("The workbook contains a server-controlled column")
    if len(set(normalized_headers)) != len(normalized_headers):
        raise UserImportParseError("The Users sheet contains duplicate headers")
    missing = [header for header in REQUIRED_HEADERS if header not in normalized_headers]
    if missing:
        raise UserImportParseError(f"The Users sheet is missing required header(s): {', '.join(missing)}")
    indexes = {header: normalized_headers.index(header) for header in REQUIRED_HEADERS}
    parsed: list[ParsedUserImportRow] = []
    for row_number, row in enumerate(rows, start=2):
        if all(_blank(value) for value in row):
            continue
        if len(parsed) >= MAX_USER_ROWS:
            raise UserImportParseError(f"The workbook may contain at most {MAX_USER_ROWS} user rows")
        values = {header: row[indexes[header]] if indexes[header] < len(row) else None for header in REQUIRED_HEADERS}
        errors: list[str] = []
        try:
            values["date_of_birth"] = _date(values["date_of_birth"])
        except ValueError as exc:
            errors.append(str(exc))
        parsed.append(ParsedUserImportRow(row_number=row_number, values=values, errors=errors))
    return parsed
